#!/usr/bin/env python3
"""
export_excel.py
===============
מייצר קובץ Excel מ-JSON של ממצאי ניתוח הסכם.

שימוש:
    python3 export_excel.py '<json_string>' <output.xlsx> <contract_name>

או:
    python3 export_excel.py findings.json output.xlsx "שם לקוח"
"""

import sys, json, os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[!] openpyxl לא מותקן. הרץ: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ─── Color palette ────────────────────────────────────────────────
DARK_GREEN   = "0F4C2A"
MID_GREEN    = "1D9E75"
LIGHT_GREEN  = "E1F5EE"
RED_BG       = "FCEBEB"
RED_TXT      = "A32D2D"
AMBER_BG     = "FAEEDA"
AMBER_TXT    = "854F0B"
OK_BG        = "EAF3DE"
OK_TXT       = "3B6D11"
HEADER_ROW   = "F5F5F5"

SEV_MAP = {"critical": "קריטי 🔴", "warning": "אזהרה 🟡", "ok": "תקין 🟢"}
RISK_MAP = {"low": "נמוך 🟢", "medium": "בינוני 🟡", "high": "גבוה 🔴"}


def hfont(color="FFFFFF", sz=11, bold=True):
    return Font(name="Arial", size=sz, bold=bold, color=color)


def cfont(sz=10, bold=False, color="000000"):
    return Font(name="Arial", size=sz, bold=bold, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def align(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, readingOrder=2)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def build_workbook(data: dict, contract_name: str) -> openpyxl.Workbook:
    findings = data.get("findings", [])
    crit = [f for f in findings if f.get("severity") == "critical"]
    warn = [f for f in findings if f.get("severity") == "warning"]
    ok   = [f for f in findings if f.get("severity") == "ok"]

    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════
    # גיליון 1 – סיכום
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "סיכום"
    ws1.sheet_view.rightToLeft = True

    # Title
    ws1.merge_cells("A1:C1")
    ws1["A1"] = f"דו\"ח ניתוח הסכם שמירה – G1 Legal AI"
    ws1["A1"].font  = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    ws1["A1"].fill  = fill(DARK_GREEN)
    ws1["A1"].alignment = align("center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:C2")
    ws1["A2"] = data.get("summary", "")
    ws1["A2"].font      = cfont(10)
    ws1["A2"].fill      = fill(LIGHT_GREEN)
    ws1["A2"].alignment = align("right", wrap=True)
    ws1.row_dimensions[2].height = 50

    stats = [
        ("", ""),
        ("נתון", "ערך"),
        ("ממצאים קריטיים", str(len(crit))),
        ("אזהרות", str(len(warn))),
        ("תקין", str(len(ok))),
        ("סה\"כ ממצאים", str(len(findings))),
        ("", ""),
        ("רמת סיכון כולל", RISK_MAP.get(data.get("overall_risk", ""), "")),
        ("", ""),
        ("הערת ביטוחים", data.get("insurance_note", "")),
        ("", ""),
        ("נספחים חסרים", data.get("missing_appendices", "")),
        ("", ""),
        ("קובץ שנבדק", contract_name),
        ("תאריך ניתוח", datetime.now().strftime("%d/%m/%Y")),
    ]

    sev_fill = {"ממצאים קריטיים": (RED_BG, RED_TXT),
                "אזהרות": (AMBER_BG, AMBER_TXT),
                "תקין": (OK_BG, OK_TXT)}

    for i, (k, v) in enumerate(stats, start=3):
        rk = ws1.cell(row=i, column=1, value=k)
        rv = ws1.cell(row=i, column=3, value=v)
        rk.alignment = align()
        rv.alignment = align("right", wrap=True)
        if k == "נתון":
            rk.font = hfont(DARK_GREEN); rk.fill = fill(LIGHT_GREEN)
            rv.font = hfont(DARK_GREEN); rv.fill = fill(LIGHT_GREEN)
        elif k in sev_fill:
            bg, tx = sev_fill[k]
            rv.font = Font(name="Arial", size=11, bold=True, color=tx)
            rv.fill = fill(bg)
        if k:
            rk.border = thin_border(); rv.border = thin_border()
        ws1.row_dimensions[i].height = 60 if k in ("הערת ביטוחים", "נספחים חסרים") else 18

    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 4
    ws1.column_dimensions["C"].width = 75

    # ═══════════════════════════════════════
    # גיליון 2 – ממצאים מפורטים
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("ממצאים מפורטים")
    ws2.sheet_view.rightToLeft = True

    headers  = ["חומרה", "כותרת ממצא", "סטטוס", "סעיף", "מצב נוכחי", "נדרש לפי הגנרי", "נוסח מוצע"]
    col_widths = [14,     30,            11,       14,      44,           44,               50]

    for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font      = hfont()
        cell.fill      = fill(DARK_GREEN)
        cell.alignment = align("center")
        cell.border    = thin_border()
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 22

    order = {"critical": 0, "warning": 1, "ok": 2}
    sorted_findings = sorted(findings, key=lambda f: order.get(f.get("severity", "ok"), 2))

    for row_i, f in enumerate(sorted_findings, start=2):
        sev = f.get("severity", "ok")
        row_bg = RED_BG if sev == "critical" else AMBER_BG if sev == "warning" else OK_BG
        sev_tx = RED_TXT if sev == "critical" else AMBER_TXT if sev == "warning" else OK_TXT
        values = [
            SEV_MAP.get(sev, sev),
            f.get("title", ""),
            f.get("tag", ""),
            f.get("clause_ref", ""),
            f.get("current", ""),
            f.get("required", ""),
            f.get("fix", ""),
        ]
        for c, val in enumerate(values, start=1):
            cell = ws2.cell(row=row_i, column=c, value=val)
            cell.alignment = align("right", wrap=True)
            cell.border    = thin_border()
            if c == 1:
                cell.font = Font(name="Arial", size=10, bold=True, color=sev_tx)
                cell.fill = fill(row_bg)
            else:
                cell.font = cfont(10)
                if sev != "ok":
                    cell.fill = fill("FEFEFE")
        ws2.row_dimensions[row_i].height = 60

    ws2.freeze_panes = "A2"

    # ═══════════════════════════════════════
    # גיליון 3 – סדר עדיפויות לתיקון
    # ═══════════════════════════════════════
    ws3 = wb.create_sheet("סדר עדיפויות לתיקון")
    ws3.sheet_view.rightToLeft = True

    ws3.merge_cells("A1:D1")
    ws3["A1"] = "סדר עדיפויות לתיקון טרום חתימה – ממצאים קריטיים"
    ws3["A1"].font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    ws3["A1"].fill      = fill(DARK_GREEN)
    ws3["A1"].alignment = align("center")
    ws3.row_dimensions[1].height = 28

    for c, (h, w) in enumerate(
        zip(["עדיפות", "ממצא", "פעולה נדרשת", "אחראי"], [10, 34, 55, 22]), start=1
    ):
        cell = ws3.cell(row=2, column=c, value=h)
        cell.font = hfont(DARK_GREEN); cell.fill = fill(LIGHT_GREEN)
        cell.alignment = align("center"); cell.border = thin_border()
        ws3.column_dimensions[get_column_letter(c)].width = w

    for i, f in enumerate(crit, start=1):
        r = i + 2
        action = (f.get("fix") or f.get("required", ""))[:250]
        bg = RED_BG if i <= 3 else AMBER_BG
        for c, val in enumerate(
            [f"קריטי #{i}", f.get("title",""), action, "לאשר עם מחלקה משפטית"], start=1
        ):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.fill = fill(bg); cell.alignment = align("right", wrap=True)
            cell.border = thin_border(); cell.font = cfont(10)
        ws3.row_dimensions[r].height = 55

    return wb


def main():
    if len(sys.argv) < 3:
        print("שימוש: python3 export_excel.py '<json>' <output.xlsx> [שם_לקוח]")
        sys.exit(1)

    json_arg      = sys.argv[1]
    output_path   = sys.argv[2]
    contract_name = sys.argv[3] if len(sys.argv) > 3 else "הסכם"

    # Accept both JSON string and file path
    if json_arg.strip().startswith("{"):
        data = json.loads(json_arg)
    elif os.path.exists(json_arg):
        with open(json_arg, encoding="utf-8") as f:
            data = json.load(f)
    else:
        print(f"[!] לא הצלחתי לפרש את הארגומנט הראשון כ-JSON או כנתיב קובץ")
        sys.exit(1)

    wb = build_workbook(data, contract_name)
    wb.save(output_path)
    print(f"✓ Excel נשמר: {output_path}")


if __name__ == "__main__":
    main()
